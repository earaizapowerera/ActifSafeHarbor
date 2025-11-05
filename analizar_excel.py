#!/usr/bin/env python3
"""
Analizar archivo Excel para identificar monedas
"""
import openpyxl
import sys

try:
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook('/Users/enrique/ActifRMF/Propuesta reporte Calculo AF.xlsx', data_only=True)

    print("=" * 80)
    print("ANÁLISIS DEL EXCEL ORIGINAL")
    print("=" * 80)

    # Analizar cada hoja
    for sheet_name in wb.sheetnames:
        print(f"\n### HOJA: {sheet_name}")
        print("=" * 80)

        ws = wb[sheet_name]

        # Leer encabezados (primera fila)
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append((cell.column, cell.value))

        print("\nCOLUMNAS:")
        for col, header in headers:
            print(f"  {col}: {header}")

        # Leer algunos valores de ejemplo (fila 2)
        print("\nVALORES DE EJEMPLO (Fila 2):")
        if ws.max_row >= 2:
            for col, header in headers:
                cell = ws.cell(row=2, column=col)
                value = cell.value
                number_format = cell.number_format

                # Identificar si tiene formato de moneda o si menciona USD/MXN
                is_currency = '$' in str(number_format) or 'currency' in str(number_format).lower()
                has_usd = 'usd' in str(header).lower() or 'dls' in str(header).lower() or 'dlls' in str(header).lower()
                has_mxn = 'mxn' in str(header).lower() or 'pesos' in str(header).lower()

                currency_indicator = ""
                if has_usd:
                    currency_indicator = " [USD]"
                elif has_mxn:
                    currency_indicator = " [MXN]"
                elif is_currency:
                    currency_indicator = " [$]"

                print(f"  {col}: {header}{currency_indicator}")
                print(f"      Valor: {value}")
                print(f"      Formato: {number_format}")

        print("\n" + "=" * 80)

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()
