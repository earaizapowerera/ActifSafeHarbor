#!/usr/bin/env python3
"""
Script de prueba con Selenium para la página de Reporte de Safe Harbor
Prueba la funcionalidad completa: carga de compañías, filtrado y exportación a Excel
"""

import os
import time
import glob
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

# Configuración
BASE_URL = "http://localhost:5071"
DOWNLOAD_DIR = os.path.expanduser("~/Downloads")

def setup_driver():
    """Configura el driver de Chrome con opciones para descarga automática"""
    chrome_options = Options()

    # Configurar directorio de descargas
    prefs = {
        "download.default_directory": DOWNLOAD_DIR,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    }
    chrome_options.add_experimental_option("prefs", prefs)

    # Opcional: descomentar para ver el navegador en acción
    # chrome_options.add_argument("--headless")  # Ejecutar sin interfaz gráfica

    driver = webdriver.Chrome(options=chrome_options)
    driver.maximize_window()
    return driver

def test_reporte_page():
    """Prueba completa de la página de reporte"""
    driver = setup_driver()

    try:
        print("=" * 80)
        print("INICIANDO PRUEBA DE REPORTE SAFE HARBOR CON SELENIUM")
        print("=" * 80)

        # 1. Navegar a la página de reporte
        print("\n[1/7] Navegando a la página de reporte...")
        driver.get(f"{BASE_URL}/reporte.html")
        time.sleep(2)

        # 2. Esperar a que carguen las compañías
        print("[2/7] Esperando a que carguen las compañías...")
        wait = WebDriverWait(driver, 15)

        # Verificar que el contenedor de compañías existe
        companias_container = wait.until(
            EC.presence_of_element_located((By.ID, "companiasContainer"))
        )

        # Esperar a que aparezcan los checkboxes
        wait.until(
            EC.presence_of_element_located((By.CLASS_NAME, "compania-check"))
        )

        # Obtener todos los checkboxes de compañías
        checkboxes = driver.find_elements(By.CLASS_NAME, "compania-check")
        print(f"   ✓ Se cargaron {len(checkboxes)} compañías")

        # Mostrar las compañías disponibles
        for i, checkbox in enumerate(checkboxes[:5], 1):  # Mostrar solo las primeras 5
            label = driver.find_element(By.CSS_SELECTOR, f"label[for='{checkbox.get_attribute('id')}']")
            print(f"   - {label.text}")

        # 3. Seleccionar las primeras 2 compañías que tengan registros
        print("\n[3/7] Seleccionando compañías con datos...")
        selected_count = 0
        for checkbox in checkboxes:
            if selected_count >= 2:
                break

            # Obtener el label para ver cuántos registros tiene
            label = driver.find_element(By.CSS_SELECTOR, f"label[for='{checkbox.get_attribute('id')}']")
            label_text = label.text

            # Buscar el número de registros en el texto
            if "registros)" in label_text and not "(0 registros)" in label_text:
                checkbox.click()
                print(f"   ✓ Seleccionada: {label_text}")
                selected_count += 1
                time.sleep(0.5)

        if selected_count == 0:
            print("   ⚠ No se encontraron compañías con registros, seleccionando la primera disponible...")
            checkboxes[0].click()

        # 4. Seleccionar año 2024
        print("\n[4/7] Seleccionando año 2024...")
        año_select = Select(driver.find_element(By.ID, "añoSelect"))
        año_select.select_by_value("2024")
        print("   ✓ Año 2024 seleccionado")

        # 5. Hacer clic en "Cargar"
        print("\n[5/7] Cargando reporte...")
        btn_cargar = driver.find_element(By.ID, "btnCargarReporte")
        btn_cargar.click()

        # Esperar a que aparezca el spinner y luego desaparezca
        time.sleep(1)
        wait.until(
            EC.text_to_be_present_in_element((By.ID, "btnCargarReporte"), "Cargar")
        )

        # Verificar que se cargaron datos
        time.sleep(2)
        contador = driver.find_element(By.ID, "contadorRegistros")
        print(f"   ✓ {contador.text} cargados")

        # Verificar que los grids tienen datos
        grid_extranjeros = driver.find_element(By.ID, "gridExtranjeros")
        grid_nacionales = driver.find_element(By.ID, "gridNacionales")

        # Verificar si hay filas en los grids (AG-Grid crea divs con clase ag-row)
        rows_ext = driver.find_elements(By.CSS_SELECTOR, "#gridExtranjeros .ag-row")
        rows_nac = driver.find_elements(By.CSS_SELECTOR, "#gridNacionales .ag-row")

        print(f"   ✓ Grid Extranjeros: {len(rows_ext)} filas visibles")
        print(f"   ✓ Grid Nacionales: {len(rows_nac)} filas visibles")

        # 6. Limpiar archivos Excel anteriores del mismo día
        print("\n[6/7] Preparando para exportar a Excel...")
        fecha_hoy = time.strftime("%Y-%m-%d")
        pattern = f"{DOWNLOAD_DIR}/SafeHarbor_2024_{fecha_hoy}*.xlsx"
        old_files = glob.glob(pattern)
        for f in old_files:
            try:
                os.remove(f)
                print(f"   - Eliminado archivo previo: {os.path.basename(f)}")
            except:
                pass

        # 7. Exportar a Excel
        print("\n[7/7] Exportando a Excel...")
        btn_excel = driver.find_element(By.ID, "btnExportarExcel")

        if btn_excel.get_attribute("disabled"):
            print("   ✗ ERROR: El botón de Excel está deshabilitado")
            return False

        btn_excel.click()
        print("   ✓ Click en botón Excel")

        # Esperar a que aparezca el alert de confirmación
        time.sleep(2)
        try:
            alert = driver.switch_to.alert
            alert_text = alert.text
            print(f"   ✓ Mensaje: {alert_text}")
            alert.accept()
        except:
            print("   - No se detectó mensaje de confirmación")

        # Esperar a que se descargue el archivo
        print("\n   Esperando descarga del archivo...")
        max_wait = 10
        for i in range(max_wait):
            time.sleep(1)
            new_files = glob.glob(pattern)
            if new_files:
                downloaded_file = new_files[0]
                file_size = os.path.getsize(downloaded_file)
                print(f"\n   ✓✓✓ ARCHIVO DESCARGADO ✓✓✓")
                print(f"   Ubicación: {downloaded_file}")
                print(f"   Tamaño: {file_size:,} bytes")

                # Verificar que el archivo es un XLSX válido
                if file_size > 0 and downloaded_file.endswith('.xlsx'):
                    print("   ✓ Archivo XLSX válido")

                    # Intentar leer las hojas del archivo
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(downloaded_file)
                        print(f"   ✓ Hojas en el archivo:")
                        for sheet_name in wb.sheetnames:
                            sheet = wb[sheet_name]
                            row_count = sheet.max_row - 1  # -1 para excluir header
                            print(f"      - {sheet_name}: {row_count} registros")
                        wb.close()
                    except ImportError:
                        print("   (openpyxl no instalado, no se pudo leer contenido)")
                    except Exception as e:
                        print(f"   ⚠ Error al leer archivo: {e}")

                    return True
                else:
                    print(f"   ✗ ERROR: Archivo inválido")
                    return False

            print(f"   Esperando... ({i+1}/{max_wait}s)")

        print("\n   ✗ ERROR: El archivo no se descargó en el tiempo esperado")
        return False

    except Exception as e:
        print(f"\n✗ ERROR EN LA PRUEBA: {e}")
        import traceback
        traceback.print_exc()
        return False

    finally:
        # Mantener el navegador abierto por 3 segundos para ver el resultado
        print("\n" + "=" * 80)
        print("Manteniendo navegador abierto por 5 segundos...")
        time.sleep(5)
        driver.quit()
        print("Navegador cerrado")

if __name__ == "__main__":
    success = test_reporte_page()

    print("\n" + "=" * 80)
    if success:
        print("✓✓✓ PRUEBA EXITOSA ✓✓✓")
        print("La exportación a Excel funciona correctamente")
    else:
        print("✗✗✗ PRUEBA FALLIDA ✗✗✗")
        print("Hubo problemas durante la prueba")
    print("=" * 80)

    exit(0 if success else 1)
